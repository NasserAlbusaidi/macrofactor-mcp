"""
MacroFactor + Garmin MCP Server — Import Layer

Handles parsing and importing MacroFactor .xlsx exports (Quick Export and
All-Time Data formats) into DuckDB.
"""

import os
import glob
import bisect
import logging
from datetime import datetime, date

import openpyxl

from schemas import init_db

logger = logging.getLogger("macrofactor-mcp")

# ═════════════════════════════════════════════════════════════════════════════
#  DATE + SHEET HELPERS
# ═════════════════════════════════════════════════════════════════════════════

WEEKDAY_MAP = {
    "Sunday": 6, "Monday": 0, "Tuesday": 1, "Wednesday": 2,
    "Thursday": 3, "Friday": 4, "Saturday": 5,
}


def _parse_date(v):
    """Convert various date formats to a YYYY-MM-DD string."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _sheet_to_dicts(ws):
    """Read a worksheet into a list of {header: value} dicts."""
    rows = list(ws.iter_rows())
    if len(rows) < 2:
        return []
    headers = [c.value for c in rows[0]]
    result = []
    for row in rows[1:]:
        if not row or all(c.value is None for c in row):
            continue
        d = {}
        for i, h in enumerate(headers):
            if h is not None and i < len(row):
                d[h] = row[i].value
        if d:
            result.append(d)
    return result


def _safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def _safe_int(v):
    if v is None:
        return None
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return None


# ═════════════════════════════════════════════════════════════════════════════
#  AUTO-IMPORT
# ═════════════════════════════════════════════════════════════════════════════

def auto_import_new_exports():
    """Auto-import any new .xlsx exports found in DATA_DIR on startup."""
    from main import get_db, DATA_DIR
    pattern = os.path.join(DATA_DIR, "*.xlsx")
    matches = sorted(glob.glob(pattern), key=os.path.getmtime, reverse=True)
    for filepath in matches:
        with get_db() as db:
            existing = db.execute(
                "SELECT filename FROM import_log WHERE filename = ?",
                [filepath],
            ).fetchone()
        if not existing:
            logger.info("Auto-importing new export: %s", filepath)
            try:
                stats = load_xlsx(filepath)
                with get_db() as db:
                    db.execute(
                        "INSERT OR REPLACE INTO import_log "
                        "(filename, format, rows_daily, rows_food, rows_workouts) "
                        "VALUES (?,?,?,?,?)",
                        [filepath, stats.get("format", "unknown"),
                         stats["daily"], stats["food"], stats["workouts"]],
                    )
                logger.info(
                    "Auto-imported (%s): %d daily, %d food, %d workout rows",
                    stats.get("format"), stats["daily"], stats["food"], stats["workouts"],
                )
            except Exception as e:
                logger.error("Auto-import failed for %s: %s", filepath, e)


# ═════════════════════════════════════════════════════════════════════════════
#  IMPORT: QUICK EXPORT FORMAT
# ═════════════════════════════════════════════════════════════════════════════

def _import_quick_export(wb, db, stats):
    """Import from the Quick Export format."""

    if "Quick Export" in wb.sheetnames:
        ws = wb["Quick Export"]
        rows = list(ws.iter_rows(min_row=1))
        if len(rows) > 1:
            headers = {
                cell.value.strip(): i for i, cell in enumerate(rows[0]) if cell.value
            }
            get = lambda r, h: (
                r[headers[h]].value if h in headers and headers[h] < len(r) else None
            )
            for row in rows[1:]:
                d = _parse_date(get(row, "Date"))
                if not d:
                    continue
                db.execute(
                    "INSERT OR REPLACE INTO daily VALUES "
                    "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    [d, get(row, "Expenditure"), get(row, "Trend Weight (kg)"),
                     get(row, "Weight (kg)"), None,
                     get(row, "Calories (kcal)"), get(row, "Protein (g)"),
                     get(row, "Fat (g)"), get(row, "Carbs (g)"),
                     get(row, "Target Calories (kcal)"), get(row, "Target Protein (g)"),
                     get(row, "Target Fat (g)"), get(row, "Target Carbs (g)"),
                     get(row, "Steps"), get(row, "Alcohol (g)"),
                     get(row, "Fiber (g)"), get(row, "Sodium (mg)"),
                     get(row, "Sugars (g)"), get(row, "Caffeine (mg)"),
                     get(row, "Calcium (mg)"), get(row, "Iron (mg)"),
                     get(row, "Vitamin C (mg)"), get(row, "Vitamin D (mcg)"),
                     get(row, "Water (g)")],
                )
                stats["daily"] += 1

    if "Food Log" in wb.sheetnames:
        ws = wb["Food Log"]
        rows = list(ws.iter_rows(min_row=1))
        if len(rows) > 1:
            headers = {
                cell.value.strip(): i for i, cell in enumerate(rows[0]) if cell.value
            }
            get = lambda r, h: (
                r[headers[h]].value if h in headers and headers[h] < len(r) else None
            )
            for row in rows[1:]:
                d = _parse_date(get(row, "Date"))
                if not d:
                    continue
                db.execute(
                    "INSERT INTO food_log VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
                    [d, str(get(row, "Time") or ""), get(row, "Food Name"),
                     get(row, "Serving Size"), get(row, "Serving Qty"),
                     get(row, "Serving Weight (g)"), get(row, "Calories (kcal)"),
                     get(row, "Protein (g)"), get(row, "Fat (g)"),
                     get(row, "Carbs (g)"), get(row, "Fiber (g)"),
                     get(row, "Sodium (mg)"), get(row, "Sugars (g)")],
                )
                stats["food"] += 1

    if "Workout Log" in wb.sheetnames:
        ws = wb["Workout Log"]
        rows = list(ws.iter_rows(min_row=1))
        if len(rows) > 1:
            headers = {
                cell.value.strip(): i for i, cell in enumerate(rows[0]) if cell.value
            }
            get = lambda r, h: (
                r[headers[h]].value if h in headers and headers[h] < len(r) else None
            )
            for row in rows[1:]:
                d = _parse_date(get(row, "Date"))
                if not d:
                    continue
                db.execute(
                    "INSERT INTO workouts VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                    [d, get(row, "Workout Duration"), get(row, "Workout"),
                     get(row, "Exercise"), get(row, "Set Type"),
                     get(row, "Weight (kg)"), get(row, "Reps"),
                     get(row, "RIR"), get(row, "Duration"),
                     get(row, "Distance short (Yd)"), get(row, "Distance long (Mi)")],
                )
                stats["workouts"] += 1

    stats["format"] = "quick"


# ═════════════════════════════════════════════════════════════════════════════
#  IMPORT: ALL-TIME DATA FORMAT
# ═════════════════════════════════════════════════════════════════════════════

def _import_alltime_export(wb, db, stats):
    """Import from the All-Time Data export format."""

    # ── Read each sheet into date-keyed dicts ────────────────────────
    macros = {}
    if "Calories & Macros" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Calories & Macros"]):
            d = _parse_date(r.get("Date"))
            if d:
                macros[d] = {
                    "calories": _safe_float(r.get("Calories (kcal)")),
                    "fat": _safe_float(r.get("Fat (g)")),
                    "carbs": _safe_float(r.get("Carbs (g)")),
                    "protein": _safe_float(r.get("Protein (g)")),
                }

    micros = {}
    if "Micronutrients" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Micronutrients"]):
            d = _parse_date(r.get("Date"))
            if d:
                micros[d] = {
                    "alcohol": _safe_float(r.get("Alcohol (g)")),
                    "fiber": _safe_float(r.get("Fiber (g)")),
                    "sodium": _safe_float(r.get("Sodium (mg)")),
                    "sugars": _safe_float(r.get("Sugars (g)")),
                    "caffeine": _safe_float(r.get("Caffeine (mg)")),
                    "calcium": _safe_float(r.get("Calcium (mg)")),
                    "iron": _safe_float(r.get("Iron (mg)")),
                    "vitamin_c": _safe_float(r.get("Vitamin C (mg)")),
                    "vitamin_d": _safe_float(r.get("Vitamin D (mcg)")),
                    "water": _safe_float(r.get("Water (g)")),
                }

    scale = {}
    if "Scale Weight" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Scale Weight"]):
            d = _parse_date(r.get("Date"))
            if d:
                scale[d] = {
                    "weight": _safe_float(r.get("Weight (kg)")),
                    "fat_pct": _safe_float(r.get("Fat Percent")),
                }

    trend = {}
    if "Weight Trend" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Weight Trend"]):
            d = _parse_date(r.get("Date"))
            if d:
                trend[d] = _safe_float(r.get("Trend Weight (kg)"))

    expenditure = {}
    if "Expenditure" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Expenditure"]):
            d = _parse_date(r.get("Date"))
            if d:
                expenditure[d] = _safe_float(r.get("Expenditure"))

    steps_data = {}
    if "Steps" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Steps"]):
            d = _parse_date(r.get("Date"))
            if d:
                steps_data[d] = _safe_int(r.get("Steps"))

    # ── Build target lookup from Nutrition Program Settings ──────────
    target_lookup = {}
    sorted_pdates = []

    if "Nutrition Program Settings" in wb.sheetnames:
        db.execute("DELETE FROM nutrition_targets")
        for r in _sheet_to_dicts(wb["Nutrition Program Settings"]):
            pd = _parse_date(r.get("Program Update Date"))
            wd_name = r.get("Program Weekday", "")
            if not pd or wd_name not in WEEKDAY_MAP:
                continue
            wd_num = WEEKDAY_MAP[wd_name]
            cal = _safe_float(r.get("Calories (kcal)"))
            fat = _safe_float(r.get("Fat (g)"))
            pro = _safe_float(r.get("Protein (g)"))
            carb = _safe_float(r.get("Carbs (g)"))

            if pd not in target_lookup:
                target_lookup[pd] = {}
            target_lookup[pd][wd_num] = (cal, pro, fat, carb)

            wd_names = {v: k for k, v in WEEKDAY_MAP.items()}
            db.execute(
                "INSERT INTO nutrition_targets VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                [pd, wd_num, wd_names.get(wd_num, ""),
                 cal, fat, pro, carb, None, None, None, None],
            )

        sorted_pdates = sorted(target_lookup.keys())

    def _resolve_targets(date_str):
        if not sorted_pdates:
            return (None, None, None, None)
        idx = bisect.bisect_right(sorted_pdates, date_str) - 1
        if idx < 0:
            return (None, None, None, None)
        pd = sorted_pdates[idx]
        wd = datetime.strptime(date_str, "%Y-%m-%d").weekday()
        return target_lookup[pd].get(wd, (None, None, None, None))

    # ── Merge all sheets by date ─────────────────────────────────────
    all_dates = set()
    all_dates.update(macros.keys())
    all_dates.update(scale.keys())
    all_dates.update(trend.keys())
    all_dates.update(expenditure.keys())
    all_dates.update(steps_data.keys())

    for d in sorted(all_dates):
        m = macros.get(d, {})
        mi = micros.get(d, {})
        s = scale.get(d, {})
        tw = trend.get(d)
        exp = expenditure.get(d)
        st = steps_data.get(d)
        tcal, tpro, tfat, tcarb = _resolve_targets(d)

        db.execute(
            "INSERT OR REPLACE INTO daily VALUES "
            "(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            [d, exp, tw, s.get("weight"), s.get("fat_pct"),
             m.get("calories"), m.get("protein"), m.get("fat"), m.get("carbs"),
             tcal, tpro, tfat, tcarb, st,
             mi.get("alcohol"), mi.get("fiber"), mi.get("sodium"),
             mi.get("sugars"), mi.get("caffeine"), mi.get("calcium"),
             mi.get("iron"), mi.get("vitamin_c"), mi.get("vitamin_d"),
             mi.get("water")],
        )
        stats["daily"] += 1

    # ── Exercise tracking sheets ─────────────────────────────────────
    exercise_sheets = {
        "Exercises - 1-RM": "estimated_1rm",
        "Exercises - 3-RM": "estimated_3rm",
        "Exercises - 10-RM": "estimated_10rm",
        "Exercises - Total Volume": "total_volume",
        "Exercises - Best Set Volume": "best_set_volume",
        "Exercises - Heaviest Weight": "heaviest_weight",
        "Exercises - Total Reps": "total_reps",
        "Exercises - Best Set Reps": "best_set_reps",
        "Exercises - Total Duration": "total_duration",
        "Exercises - Best Set Duration": "best_set_duration",
        "Exercises - Total Distance": "total_distance",
        "Exercises - Best Set Distance": "best_set_distance",
        "Exercises - Total Sets": "total_sets",
    }

    for sheet_name, metric in exercise_sheets.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows())
        if len(rows) < 2:
            continue
        headers = [c.value for c in rows[0]]
        for row in rows[1:]:
            if not row or len(row) == 0:
                continue
            d = _parse_date(row[0].value)
            if not d:
                continue
            for i, h in enumerate(headers[1:], 1):
                if h and i < len(row) and row[i].value is not None:
                    exercise = h
                    for suffix in (" (kg)", " (sets)", " (reps)", " (sec)", " (mi)", " (yd)"):
                        if exercise.endswith(suffix):
                            exercise = exercise[: -len(suffix)]
                            break
                    val = _safe_float(row[i].value)
                    if val is not None:
                        db.execute(
                            "INSERT INTO exercise_tracking VALUES (?,?,?,?)",
                            [d, exercise, metric, val],
                        )

    # ── Body Metrics ─────────────────────────────────────────────────
    if "Body Metrics" in wb.sheetnames:
        ws = wb["Body Metrics"]
        rows = list(ws.iter_rows())
        if len(rows) > 1:
            headers = [c.value for c in rows[0]]
            for row in rows[1:]:
                if not row or len(row) == 0:
                    continue
                d = _parse_date(row[0].value)
                if not d:
                    continue
                for i, h in enumerate(headers[1:], 1):
                    if h and i < len(row) and row[i].value is not None:
                        metric_name = h.replace(" (in)", "").strip()
                        val = _safe_float(row[i].value)
                        if val is not None:
                            db.execute(
                                "INSERT INTO body_metrics VALUES (?,?,?)",
                                [d, metric_name, val],
                            )

    # ── Custom Foods ─────────────────────────────────────────────────
    if "Custom Foods" in wb.sheetnames:
        db.execute("DELETE FROM custom_foods")
        for r in _sheet_to_dicts(wb["Custom Foods"]):
            name = r.get("Food Name")
            if not name:
                continue
            db.execute(
                "INSERT INTO custom_foods VALUES (?,?,?,?,?,?,?,?,?,?,?)",
                [name, r.get("Serving Size"),
                 _safe_float(r.get("Serving Qty")),
                 _safe_float(r.get("Serving Weight (g)")),
                 _safe_float(r.get("Calories (kcal)")),
                 _safe_float(r.get("Protein (g)")),
                 _safe_float(r.get("Fat (g)")),
                 _safe_float(r.get("Carbs (g)")),
                 _safe_float(r.get("Fiber (g)")),
                 _safe_float(r.get("Sodium (mg)")),
                 _safe_float(r.get("Sugars (g)"))],
            )

    # ── Food Log Notes ───────────────────────────────────────────────
    if "Food Log Notes" in wb.sheetnames:
        for r in _sheet_to_dicts(wb["Food Log Notes"]):
            d = _parse_date(r.get("Date"))
            note = r.get("Food Log Notes")
            if d and note:
                db.execute(
                    "INSERT INTO food_log_notes VALUES (?,?,?)",
                    [d, r.get("Name", ""), str(note)],
                )

    stats["format"] = "alltime"


# ═════════════════════════════════════════════════════════════════════════════
#  IMPORT: MUSCLE GROUPS (shared by both formats)
# ═════════════════════════════════════════════════════════════════════════════

def _import_muscle_groups(wb, db):
    for sheet_name, table in [
        ("Muscle Groups - Sets", "muscle_sets"),
        ("Muscle Groups - Volume", "muscle_volume"),
    ]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1))
        if len(rows) > 1:
            header_cells = [cell.value for cell in rows[0]]
            for row in rows[1:]:
                if not row or len(row) == 0:
                    continue
                d = _parse_date(row[0].value)
                if not d:
                    continue
                for i, h in enumerate(header_cells[1:], 1):
                    if h and i < len(row) and row[i].value is not None:
                        muscle = (
                            h.replace(" (sets)", "")
                            .replace(" (kg)", "")
                            .strip()
                        )
                        db.execute(
                            f"INSERT INTO {table} VALUES (?,?,?)",
                            [d, muscle, row[i].value],
                        )


# ═════════════════════════════════════════════════════════════════════════════
#  MAIN IMPORT DISPATCHER
# ═════════════════════════════════════════════════════════════════════════════

def load_xlsx(filepath: str) -> dict:
    """Load a MacroFactor .xlsx export into DuckDB. Returns row counts."""
    from main import get_db
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    stats = {"daily": 0, "food": 0, "workouts": 0, "format": "unknown"}

    with get_db() as db:
        init_db(db)

        if "Quick Export" in wb.sheetnames:
            logger.info("Detected Quick Export format")
            _import_quick_export(wb, db, stats)
        elif "Calories & Macros" in wb.sheetnames:
            logger.info("Detected All-Time Data format")
            _import_alltime_export(wb, db, stats)
        else:
            logger.warning("Unknown export format. Sheets: %s", wb.sheetnames)

        _import_muscle_groups(wb, db)

    wb.close()
    return stats
